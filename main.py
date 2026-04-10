"""
Excel多表合并工具
- 选择文件夹，自动读取所有 .xlsx/.xls 文件并合并
- 支持子文件夹扫描、来源列、多种合并模式
- .xlsx 文件完整保留原始单元格样式（填充色、字体颜色等）
- tkinter GUI，中文界面
"""

import copy
import os
import platform
import subprocess
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── 颜色 / 字体常量 ──────────────────────────────────────────────
BG       = "#F7F8FA"
CARD_BG  = "#FFFFFF"
PRIMARY  = "#4A90D9"
TEXT     = "#333333"
SUB_TEXT = "#888888"
FONT     = ("Microsoft YaHei UI", 10)
FONT_B   = ("Microsoft YaHei UI", 10, "bold")
FONT_S   = ("Microsoft YaHei UI", 9)
FONT_T   = ("Microsoft YaHei UI", 14, "bold")

# 来源列表头样式
_SOURCE_HEADER_FONT  = Font(name="Microsoft YaHei UI", size=11, bold=True, color="FFFFFF")
_SOURCE_HEADER_FILL  = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
_SOURCE_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


class ExcelMergerApp:
    """主窗口."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel多表合并工具")
        self.root.geometry("620x700")
        self.root.resizable(False, True)
        self.root.configure(bg=BG)

        self.folder_path        = tk.StringVar()
        self.include_subfolders = tk.BooleanVar(value=False)
        self.add_source_col     = tk.BooleanVar(value=True)
        self.merge_mode         = tk.StringVar(value="single")
        self.sheet_scope        = tk.StringVar(value="first")

        self._build_ui()

    # ── UI 构建 ───────────────────────────────────────────────────
    def _build_ui(self):
        tk.Label(self.root, text="Excel 多表合并工具", font=FONT_T,
                 bg=BG, fg=PRIMARY).pack(pady=(18, 6))
        tk.Label(self.root, text="选择文件夹，一键合并所有 Excel 文件", font=FONT_S,
                 bg=BG, fg=SUB_TEXT).pack(pady=(0, 12))

        # ── 文件夹选择卡片 ──
        card1 = self._card(self.root)
        card1.pack(fill="x", padx=24, pady=(0, 8))
        row = tk.Frame(card1, bg=CARD_BG)
        row.pack(fill="x", padx=12, pady=10)
        tk.Label(row, text="文件夹路径：", font=FONT_B, bg=CARD_BG, fg=TEXT).pack(side="left")
        tk.Entry(row, textvariable=self.folder_path, font=FONT, width=34,
                 relief="solid", bd=1).pack(side="left", padx=(4, 8))
        tk.Button(row, text="浏 览", font=FONT_S, bg=PRIMARY, fg="white",
                  activebackground="#3A7BC8", activeforeground="white",
                  relief="flat", padx=12, pady=2, cursor="hand2",
                  command=self._browse_folder).pack(side="left")

        # ── 选项卡片 ──
        card2 = self._card(self.root)
        card2.pack(fill="x", padx=24, pady=(0, 8))
        tk.Label(card2, text="合并选项", font=FONT_B, bg=CARD_BG, fg=TEXT).pack(
            anchor="w", padx=14, pady=(10, 4))
        ttk.Separator(card2, orient="horizontal").pack(fill="x", padx=14)
        opts = tk.Frame(card2, bg=CARD_BG)
        opts.pack(fill="x", padx=14, pady=8)

        chk_row1 = tk.Frame(opts, bg=CARD_BG)
        chk_row1.pack(fill="x", pady=2)
        tk.Checkbutton(chk_row1, text="包含子文件夹", variable=self.include_subfolders,
                       font=FONT, bg=CARD_BG, fg=TEXT, activebackground=CARD_BG,
                       selectcolor=CARD_BG).pack(side="left", padx=(0, 24))
        tk.Checkbutton(chk_row1, text="添加「来源文件名」列", variable=self.add_source_col,
                       font=FONT, bg=CARD_BG, fg=TEXT, activebackground=CARD_BG,
                       selectcolor=CARD_BG).pack(side="left")

        mode_frame = tk.LabelFrame(opts, text="合并模式", font=FONT_S,
                                   bg=CARD_BG, fg=SUB_TEXT, bd=1, relief="groove")
        mode_frame.pack(fill="x", pady=(8, 4))
        for text, val in [("合并到同一个 Sheet", "single"), ("每个文件单独一个 Sheet", "separate")]:
            tk.Radiobutton(mode_frame, text=text, variable=self.merge_mode, value=val,
                           font=FONT, bg=CARD_BG, fg=TEXT, activebackground=CARD_BG,
                           selectcolor=CARD_BG).pack(side="left", padx=(8, 20), pady=4)

        scope_frame = tk.LabelFrame(opts, text="读取范围", font=FONT_S,
                                    bg=CARD_BG, fg=SUB_TEXT, bd=1, relief="groove")
        scope_frame.pack(fill="x", pady=(4, 4))
        for text, val in [("只读取第一个 Sheet", "first"), ("读取所有 Sheet", "all")]:
            tk.Radiobutton(scope_frame, text=text, variable=self.sheet_scope, value=val,
                           font=FONT, bg=CARD_BG, fg=TEXT, activebackground=CARD_BG,
                           selectcolor=CARD_BG).pack(side="left", padx=(8, 20), pady=4)

        # ── 进度条 ──
        card3 = self._card(self.root)
        card3.pack(fill="x", padx=24, pady=(0, 8))
        prog_frame = tk.Frame(card3, bg=CARD_BG)
        prog_frame.pack(fill="x", padx=14, pady=10)
        self.progress = ttk.Progressbar(prog_frame, orient="horizontal",
                                        length=540, mode="determinate")
        self.progress.pack(fill="x")
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(prog_frame, textvariable=self.status_var, font=FONT_S,
                 bg=CARD_BG, fg=SUB_TEXT, anchor="w").pack(fill="x", pady=(4, 0))

        # ── 日志 ──
        card4 = self._card(self.root)
        card4.pack(fill="x", padx=24, pady=(0, 8))
        self.log_text = tk.Text(card4, height=5, font=("Consolas", 9), bg="#FAFAFA",
                                fg=TEXT, relief="flat", wrap="word", state="disabled")
        self.log_text.pack(fill="x", padx=10, pady=6)

        # ── 执行按钮 ──
        btn_frame = tk.Frame(self.root, bg=BG)
        btn_frame.pack(fill="x", padx=24, pady=(8, 16))
        tk.Button(btn_frame, text="开 始 合 并",
                  font=("Microsoft YaHei UI", 13, "bold"),
                  bg=PRIMARY, fg="white", activebackground="#3A7BC8",
                  activeforeground="white", relief="flat",
                  padx=60, pady=10, cursor="hand2",
                  command=self._start_merge).pack(expand=True)

    @staticmethod
    def _card(parent) -> tk.Frame:
        return tk.Frame(parent, bg=CARD_BG, bd=1, relief="solid", highlightthickness=0)

    def _log(self, msg: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _browse_folder(self):
        path = filedialog.askdirectory(title="选择包含 Excel 文件的文件夹")
        if path:
            self.folder_path.set(path)

    # ── 启动合并 ──────────────────────────────────────────────────
    def _start_merge(self):
        folder = self.folder_path.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("提示", "请先选择一个有效的文件夹路径。")
            return
        self._clear_log()
        self.progress["value"] = 0
        self.status_var.set("扫描文件中…")
        threading.Thread(target=self._merge_worker, args=(folder,), daemon=True).start()

    # ── 合并核心 ──────────────────────────────────────────────────
    def _merge_worker(self, folder: str):
        try:
            files = self._scan_files(folder)
            if not files:
                self.root.after(0, lambda: messagebox.showinfo(
                    "提示", "所选文件夹内未找到 .xlsx 或 .xls 文件。"))
                self.root.after(0, lambda: self.status_var.set("就绪"))
                return

            total = len(files)
            self.root.after(0, lambda: self._log(f"共找到 {total} 个 Excel 文件"))

            merge_mode  = self.merge_mode.get()
            sheet_scope = self.sheet_scope.get()
            add_source  = self.add_source_col.get()
            skipped: list[str] = []

            out_wb = openpyxl.Workbook()
            out_wb.remove(out_wb.active)  # 删除默认空 sheet

            if merge_mode == "single":
                out_ws = out_wb.create_sheet("合并结果")
                current_row   = 1
                header_written = False

                for idx, fpath in enumerate(files, 1):
                    fname = os.path.basename(fpath)
                    self.root.after(0, lambda f=fname, i=idx:
                                    self.status_var.set(f"处理中 ({i}/{total})：{f}"))
                    try:
                        sheets = self._read_file(fpath, sheet_scope)
                    except Exception as e:
                        skipped.append(fname)
                        self.root.after(0, lambda f=fname, err=e:
                                        self._log(f"[跳过] {f}：{err}"))
                        self._update_progress(idx, total)
                        continue

                    if not sheets:
                        skipped.append(fname)
                        self.root.after(0, lambda f=fname:
                                        self._log(f"[跳过] {f}：文件无有效数据"))
                        self._update_progress(idx, total)
                        continue

                    show_sheet_col = add_source and sheet_scope == "all" and len(sheets) > 1

                    for sheet_name, rows_data, is_df in sheets:
                        if is_df:
                            df = rows_data
                            if not header_written:
                                col = 1
                                if add_source:
                                    self._write_source_header(out_ws, current_row, col, "来源文件名")
                                    col += 1
                                if show_sheet_col:
                                    self._write_source_header(out_ws, current_row, col, "来源Sheet")
                                    col += 1
                                for h in df.columns:
                                    out_ws.cell(row=current_row, column=col, value=str(h))
                                    col += 1
                                current_row += 1
                                header_written = True
                            for _, data_row in df.iterrows():
                                col = 1
                                if add_source:
                                    out_ws.cell(row=current_row, column=col, value=fname)
                                    col += 1
                                if show_sheet_col:
                                    out_ws.cell(row=current_row, column=col, value=sheet_name)
                                    col += 1
                                for val in data_row:
                                    out_ws.cell(row=current_row, column=col, value=val)
                                    col += 1
                                current_row += 1
                        else:
                            # rows_data = list of rows, each row = list of cell dicts
                            if not rows_data:
                                continue
                            if not header_written:
                                col = 1
                                if add_source:
                                    self._write_source_header(out_ws, current_row, col, "来源文件名")
                                    col += 1
                                if show_sheet_col:
                                    self._write_source_header(out_ws, current_row, col, "来源Sheet")
                                    col += 1
                                for cd in rows_data[0]:
                                    self._apply_cell(out_ws.cell(row=current_row, column=col), cd)
                                    col += 1
                                current_row += 1
                                header_written = True
                                data_rows = rows_data[1:]
                            else:
                                data_rows = rows_data[1:]

                            for src_row in data_rows:
                                col = 1
                                if add_source:
                                    out_ws.cell(row=current_row, column=col, value=fname)
                                    col += 1
                                if show_sheet_col:
                                    out_ws.cell(row=current_row, column=col, value=sheet_name)
                                    col += 1
                                for cd in src_row:
                                    self._apply_cell(out_ws.cell(row=current_row, column=col), cd)
                                    col += 1
                                current_row += 1

                    self.root.after(0, lambda f=fname: self._log(f"[完成] {f}"))
                    self._update_progress(idx, total)

                if current_row == 1:
                    self.root.after(0, lambda: messagebox.showinfo("提示", "没有可合并的数据。"))
                    self.root.after(0, lambda: self.status_var.set("就绪"))
                    return

                self._auto_fit(out_ws)
                out_ws.freeze_panes = "A2"

            else:  # separate
                used_names: set[str] = set()

                for idx, fpath in enumerate(files, 1):
                    fname = os.path.basename(fpath)
                    self.root.after(0, lambda f=fname, i=idx:
                                    self.status_var.set(f"处理中 ({i}/{total})：{f}"))
                    try:
                        sheets = self._read_file(fpath, sheet_scope)
                    except Exception as e:
                        skipped.append(fname)
                        self.root.after(0, lambda f=fname, err=e:
                                        self._log(f"[跳过] {f}：{err}"))
                        self._update_progress(idx, total)
                        continue

                    if not sheets:
                        skipped.append(fname)
                        self.root.after(0, lambda f=fname:
                                        self._log(f"[跳过] {f}：文件无有效数据"))
                        self._update_progress(idx, total)
                        continue

                    for sheet_name, rows_data, is_df in sheets:
                        base = Path(fname).stem
                        key  = base if len(sheets) == 1 else f"{base}_{sheet_name}"
                        key  = key[:31]
                        orig = key
                        cnt  = 1
                        while key in used_names:
                            sfx = f"_{cnt}"
                            key = orig[:31 - len(sfx)] + sfx
                            cnt += 1
                        used_names.add(key)

                        out_ws = out_wb.create_sheet(key)
                        row = 1

                        if is_df:
                            df = rows_data
                            col = 1
                            if add_source:
                                self._write_source_header(out_ws, row, col, "来源文件名")
                                col += 1
                            for h in df.columns:
                                out_ws.cell(row=row, column=col, value=str(h))
                                col += 1
                            row += 1
                            for _, data_row in df.iterrows():
                                col = 1
                                if add_source:
                                    out_ws.cell(row=row, column=col, value=fname)
                                    col += 1
                                for val in data_row:
                                    out_ws.cell(row=row, column=col, value=val)
                                    col += 1
                                row += 1
                        else:
                            for r_idx, src_row in enumerate(rows_data):
                                col = 1
                                if add_source:
                                    if r_idx == 0:
                                        self._write_source_header(out_ws, row, col, "来源文件名")
                                    else:
                                        out_ws.cell(row=row, column=col, value=fname)
                                    col += 1
                                for cd in src_row:
                                    self._apply_cell(out_ws.cell(row=row, column=col), cd)
                                    col += 1
                                row += 1

                        self._auto_fit(out_ws)
                        out_ws.freeze_panes = "A2"

                    self.root.after(0, lambda f=fname: self._log(f"[完成] {f}"))
                    self._update_progress(idx, total)

                if not out_wb.worksheets:
                    self.root.after(0, lambda: messagebox.showinfo("提示", "没有可合并的数据。"))
                    self.root.after(0, lambda: self.status_var.set("就绪"))
                    return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_name  = f"合并结果_{timestamp}.xlsx"
            out_path  = os.path.join(folder, out_name)
            self.root.after(0, lambda: self.status_var.set("正在写入文件…"))
            out_wb.save(out_path)

            summary = f"合并完成！输出文件：{out_name}"
            if skipped:
                summary += f"\n\n以下 {len(skipped)} 个文件被跳过：\n" + "\n".join(skipped)

            self.root.after(0, lambda: self.status_var.set("完成"))
            self.root.after(0, lambda: self._log(f"输出文件：{out_path}"))
            self.root.after(0, lambda s=summary: messagebox.showinfo("完成", s))
            self.root.after(0, lambda: self._open_folder(os.path.dirname(out_path)))

        except Exception as e:
            self.root.after(0, lambda err=e: messagebox.showerror(
                "错误", f"合并过程中发生错误：\n{err}"))
            self.root.after(0, lambda: self.status_var.set("出错"))

    # ── 读取文件 ──────────────────────────────────────────────────
    def _read_file(self, fpath: str, scope: str) -> list[tuple]:
        """返回 list of (sheet_name, data, is_df)。
        .xlsx → data 为行列式列表（保留样式）；
        .xls  → data 为 DataFrame（xlrd 不支持样式）。
        """
        if fpath.lower().endswith(".xls"):
            if scope == "first":
                df = pd.read_excel(fpath, sheet_name=0, engine="xlrd")
                xl = pd.ExcelFile(fpath, engine="xlrd")
                sname = xl.sheet_names[0]
                xl.close()
                return [(sname, df, True)] if not df.empty else []
            else:
                dfs = pd.read_excel(fpath, sheet_name=None, engine="xlrd")
                return [(n, df, True) for n, df in dfs.items() if not df.empty]

        # .xlsx：用 openpyxl 直接读，保留样式
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws_list = [wb.worksheets[0]] if scope == "first" else list(wb.worksheets)
        results = []
        for ws in ws_list:
            if not ws.max_row:
                continue
            rows_data = []
            for row in ws.iter_rows():
                row_cells = []
                for cell in row:
                    row_cells.append({
                        "value":         cell.value,
                        "font":          copy.copy(cell.font)      if cell.has_style else None,
                        "fill":          copy.copy(cell.fill)      if cell.has_style else None,
                        "border":        copy.copy(cell.border)    if cell.has_style else None,
                        "alignment":     copy.copy(cell.alignment) if cell.has_style else None,
                        "number_format": cell.number_format,
                    })
                rows_data.append(row_cells)
            if rows_data:
                results.append((ws.title, rows_data, False))
        wb.close()
        return results

    # ── 单元格写入 ────────────────────────────────────────────────
    @staticmethod
    def _apply_cell(dst, cd: dict):
        """把带样式的 cell dict 写入目标单元格，完整保留原始格式。"""
        dst.value = cd["value"]
        if cd.get("font"):
            dst.font = cd["font"]
        if cd.get("fill"):
            dst.fill = cd["fill"]
        if cd.get("border"):
            dst.border = cd["border"]
        if cd.get("alignment"):
            dst.alignment = cd["alignment"]
        if cd.get("number_format"):
            dst.number_format = cd["number_format"]

    @staticmethod
    def _write_source_header(ws, row: int, col: int, text: str):
        """写入来源列表头（蓝底白字）。"""
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = _SOURCE_HEADER_FONT
        cell.fill      = _SOURCE_HEADER_FILL
        cell.alignment = _SOURCE_HEADER_ALIGN

    # ── 自动列宽 ──────────────────────────────────────────────────
    @staticmethod
    def _auto_fit(ws):
        for col_idx in range(1, (ws.max_column or 0) + 1):
            max_len  = 0
            col_ltr  = get_column_letter(col_idx)
            max_row  = min(ws.max_row or 1, 300)
            for row in ws.iter_rows(min_row=1, max_row=max_row,
                                    min_col=col_idx, max_col=col_idx):
                val = row[0].value
                if val is not None:
                    length = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    if length > max_len:
                        max_len = length
            ws.column_dimensions[col_ltr].width = min(max(max_len + 3, 8), 55)

    # ── 扫描文件 ──────────────────────────────────────────────────
    def _scan_files(self, folder: str) -> list[str]:
        result = []
        if self.include_subfolders.get():
            for root_dir, _, filenames in os.walk(folder):
                for fn in filenames:
                    if fn.lower().endswith((".xlsx", ".xls")) and not fn.startswith("~$"):
                        result.append(os.path.join(root_dir, fn))
        else:
            for fn in os.listdir(folder):
                if fn.lower().endswith((".xlsx", ".xls")) and not fn.startswith("~$"):
                    full = os.path.join(folder, fn)
                    if os.path.isfile(full):
                        result.append(full)
        result.sort()
        return result

    # ── 进度 ──────────────────────────────────────────────────────
    def _update_progress(self, current: int, total: int):
        pct = current / total * 100
        self.root.after(0, lambda: self.progress.configure(value=pct))

    # ── 打开文件夹 ────────────────────────────────────────────────
    @staticmethod
    def _open_folder(path: str):
        system = platform.system()
        if system == "Windows":
            os.startfile(path)
        elif system == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])


def main():
    try:
        if platform.system() == "Windows":
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    root = tk.Tk()
    ExcelMergerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()